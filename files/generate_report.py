#!/usr/bin/env python3
"""
Subnet Sweep Report Generator
Consolidates subnet sweep results into CSV/Excel format
"""

import os
import re
import argparse
from datetime import datetime
import csv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[!] Warning: openpyxl not installed. Excel export disabled.")
    print("[!] Install with: pip install openpyxl --break-system-packages")


class ReportGenerator:
    def __init__(self, results_dir, output_file="subnet_report", format_type="both"):
        self.results_dir = results_dir
        self.output_file = output_file
        self.format_type = format_type
        self.subnet_data = {}
    
    def parse_results_directory(self):
        """Parse all result files in the directory"""
        print(f"[*] Scanning directory: {self.results_dir}")
        
        if not os.path.exists(self.results_dir):
            print(f"[!] Error: Directory '{self.results_dir}' not found")
            return False
        
        # Find all IP-only files (they have clean IP lists)
        ip_files = []
        for filename in os.listdir(self.results_dir):
            if filename.endswith('_ips_only.txt'):
                ip_files.append(filename)
        
        if not ip_files:
            print(f"[!] No result files found in '{self.results_dir}'")
            return False
        
        print(f"[*] Found {len(ip_files)} subnet result file(s)")
        
        # Parse each file
        for filename in ip_files:
            filepath = os.path.join(self.results_dir, filename)
            subnet = self.extract_subnet_from_filename(filename)
            
            if subnet:
                ips = self.read_ips_from_file(filepath)
                if ips:
                    # Also try to get detailed info from the detailed file
                    detailed_file = filename.replace('_ips_only.txt', '.txt')
                    detailed_path = os.path.join(self.results_dir, detailed_file)
                    
                    if os.path.exists(detailed_path):
                        ip_details = self.parse_detailed_file(detailed_path)
                        self.subnet_data[subnet] = {
                            'ips': ips,
                            'count': len(ips),
                            'details': ip_details
                        }
                    else:
                        self.subnet_data[subnet] = {
                            'ips': ips,
                            'count': len(ips),
                            'details': {}
                        }
                    
                    print(f"[+] {subnet}: {len(ips)} alive host(s)")
        
        return len(self.subnet_data) > 0
    
    def extract_subnet_from_filename(self, filename):
        """Extract subnet from filename format: 192.168.1.0_24_timestamp_ips_only.txt"""
        # Remove _ips_only.txt and timestamp
        pattern = r'^(.+?)_(\d{8}_\d{6})_ips_only\.txt$'
        match = re.match(pattern, filename)
        
        if match:
            subnet_part = match.group(1)
            # Convert underscores back to dots and slashes
            # e.g., 192.168.1.0_24 -> 192.168.1.0/24
            subnet_part = subnet_part.replace('_', '.', 3)  # Replace first 3 underscores with dots
            subnet_part = subnet_part.replace('.', '/', 1)[::-1].replace('.', '/', 1)[::-1]  # Last underscore to slash
            
            # Better approach: use regex
            subnet_match = re.match(r'([\d\.]+)_(\d+)', match.group(1))
            if subnet_match:
                return f"{subnet_match.group(1)}/{subnet_match.group(2)}"
        
        return None
    
    def read_ips_from_file(self, filepath):
        """Read IPs from file"""
        try:
            with open(filepath, 'r') as f:
                ips = [line.strip() for line in f if line.strip()]
            return ips
        except Exception as e:
            print(f"[!] Error reading {filepath}: {e}")
            return []
    
    def parse_detailed_file(self, filepath):
        """Parse detailed results file to get detection methods"""
        ip_details = {}
        try:
            with open(filepath, 'r') as f:
                content = f.read()
                
            # Find the section with IP details
            lines = content.split('\n')
            in_data_section = False
            
            for line in lines:
                if '----' in line and in_data_section:
                    continue
                if 'IP Address' in line and 'Detection Methods' in line:
                    in_data_section = True
                    continue
                
                if in_data_section and line.strip():
                    parts = line.split('\t')
                    if len(parts) >= 2:
                        ip = parts[0].strip()
                        methods = '\t'.join(parts[1:]).strip()
                        if ip and methods:
                            ip_details[ip] = methods
            
            return ip_details
        except Exception as e:
            print(f"[!] Error parsing detailed file {filepath}: {e}")
            return {}
    
    def generate_csv(self):
        """Generate CSV report"""
        csv_filename = f"{self.output_file}.csv"
        
        try:
            with open(csv_filename, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                writer.writerow(['Subnet', 'Total Alive IPs', 'IP Addresses', 'Detection Methods'])
                
                # Write data
                for subnet in sorted(self.subnet_data.keys()):
                    data = self.subnet_data[subnet]
                    ips_list = data['ips']
                    count = data['count']
                    
                    # Create IP list with detection methods
                    ip_details_list = []
                    for ip in ips_list:
                        methods = data['details'].get(ip, 'N/A')
                        ip_details_list.append(f"{ip} ({methods})")
                    
                    # Join IPs with newline character (will appear as separate lines in Excel)
                    ips_str = '\n'.join(ips_list)
                    details_str = '\n'.join(ip_details_list)
                    
                    writer.writerow([subnet, count, ips_str, details_str])
            
            print(f"\n[+] CSV report generated: {csv_filename}")
            return csv_filename
        
        except Exception as e:
            print(f"[!] Error generating CSV: {e}")
            return None
    
    def generate_excel(self):
        """Generate Excel report with formatting"""
        if not EXCEL_AVAILABLE:
            print("[!] Cannot generate Excel file - openpyxl not installed")
            return None
        
        excel_filename = f"{self.output_file}.xlsx"
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Subnet Sweep Results"
            
            # Define styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            cell_alignment = Alignment(vertical="top", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Set column widths
            ws.column_dimensions['A'].width = 20  # Subnet
            ws.column_dimensions['B'].width = 15  # Count
            ws.column_dimensions['C'].width = 20  # IPs
            ws.column_dimensions['D'].width = 50  # Detection Methods
            
            # Write header
            headers = ['Subnet', 'Total Alive IPs', 'IP Addresses', 'Detection Methods']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            ws.row_dimensions[1].height = 30
            
            # Write data
            row = 2
            for subnet in sorted(self.subnet_data.keys()):
                data = self.subnet_data[subnet]
                ips_list = data['ips']
                count = data['count']
                
                # Create detailed IP list
                ip_details_list = []
                for ip in ips_list:
                    methods = data['details'].get(ip, 'N/A')
                    ip_details_list.append(f"{ip} - {methods}")
                
                # Subnet
                cell_a = ws.cell(row=row, column=1)
                cell_a.value = subnet
                cell_a.alignment = cell_alignment
                cell_a.border = border
                
                # Count
                cell_b = ws.cell(row=row, column=2)
                cell_b.value = count
                cell_b.alignment = Alignment(horizontal="center", vertical="center")
                cell_b.border = border
                
                # IPs
                cell_c = ws.cell(row=row, column=3)
                cell_c.value = '\n'.join(ips_list)
                cell_c.alignment = cell_alignment
                cell_c.border = border
                
                # Detection Methods
                cell_d = ws.cell(row=row, column=4)
                cell_d.value = '\n'.join(ip_details_list)
                cell_d.alignment = cell_alignment
                cell_d.border = border
                
                # Set row height based on number of IPs
                ws.row_dimensions[row].height = max(15 * len(ips_list), 20)
                
                row += 1
            
            # Add summary sheet
            ws_summary = wb.create_sheet("Summary")
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 20
            
            # Summary header
            ws_summary['A1'] = "Metric"
            ws_summary['B1'] = "Value"
            ws_summary['A1'].font = header_font
            ws_summary['B1'].font = header_font
            ws_summary['A1'].fill = header_fill
            ws_summary['B1'].fill = header_fill
            ws_summary['A1'].alignment = header_alignment
            ws_summary['B1'].alignment = header_alignment
            
            # Summary data
            total_subnets = len(self.subnet_data)
            total_ips = sum(data['count'] for data in self.subnet_data.values())
            
            summary_data = [
                ("Total Subnets Scanned", total_subnets),
                ("Total Alive IPs Discovered", total_ips),
                ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("Average IPs per Subnet", f"{total_ips/total_subnets:.2f}" if total_subnets > 0 else "0")
            ]
            
            for idx, (metric, value) in enumerate(summary_data, 2):
                ws_summary[f'A{idx}'] = metric
                ws_summary[f'B{idx}'] = value
                ws_summary[f'A{idx}'].border = border
                ws_summary[f'B{idx}'].border = border
            
            # Save workbook
            wb.save(excel_filename)
            print(f"[+] Excel report generated: {excel_filename}")
            return excel_filename
        
        except Exception as e:
            print(f"[!] Error generating Excel: {e}")
            return None
    
    def generate_reports(self):
        """Generate reports based on format type"""
        print(f"\n{'='*60}")
        print("Generating Consolidated Report")
        print(f"{'='*60}\n")
        
        if not self.parse_results_directory():
            return False
        
        if not self.subnet_data:
            print("[!] No data to generate report")
            return False
        
        print(f"\n[*] Summary:")
        print(f"    - Total Subnets: {len(self.subnet_data)}")
        print(f"    - Total Alive IPs: {sum(d['count'] for d in self.subnet_data.values())}")
        print(f"\n[*] Generating report(s)...\n")
        
        generated_files = []
        
        if self.format_type in ['csv', 'both']:
            csv_file = self.generate_csv()
            if csv_file:
                generated_files.append(csv_file)
        
        if self.format_type in ['excel', 'both']:
            excel_file = self.generate_excel()
            if excel_file:
                generated_files.append(excel_file)
        
        if generated_files:
            print(f"\n{'='*60}")
            print("Report Generation Complete!")
            print(f"{'='*60}")
            print("\nGenerated files:")
            for f in generated_files:
                print(f"  - {f}")
            print()
            return True
        else:
            print("[!] Failed to generate any reports")
            return False


def main():
    parser = argparse.ArgumentParser(
        description='Generate consolidated CSV/Excel report from subnet sweep results',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 generate_report.py -d scan_results
  python3 generate_report.py -d scan_results -o my_report -f excel
  python3 generate_report.py -d scan_results -o final_report -f both

Format Options:
  csv   - Generate CSV file only
  excel - Generate Excel file only (requires openpyxl)
  both  - Generate both CSV and Excel (default)

Report Contents:
  - Subnet-wise summary
  - Total alive IPs per subnet
  - Individual IP addresses
  - Detection methods used
  - Summary statistics
        """
    )
    
    parser.add_argument('-d', '--directory', 
                       default='scan_results',
                       help='Directory containing scan results (default: scan_results)')
    parser.add_argument('-o', '--output', 
                       default='subnet_sweep_report',
                       help='Output filename without extension (default: subnet_sweep_report)')
    parser.add_argument('-f', '--format', 
                       choices=['csv', 'excel', 'both'],
                       default='both',
                       help='Output format (default: both)')
    
    args = parser.parse_args()
    
    # Generate reports
    generator = ReportGenerator(args.directory, args.output, args.format)
    generator.generate_reports()


if __name__ == "__main__":
    main()
